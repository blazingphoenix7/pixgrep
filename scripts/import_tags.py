"""CLI: import tags from xlsx/csv into the pixgrep index.

Usage:
    python scripts/import_tags.py --file data.xlsx --sheet Sheet1 --mapping _local/tags_mapping.json

Mapping JSON format:
    {
        "filename": "<col>",
        "fields": {"canonical_name": "<col>", ...},
        "text": ["<col>", ...]
    }
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from pixgrep.config import load_config
from pixgrep.tags import import_tags


def read_xlsx(path: Path, sheet: str | None) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        rec: dict = {}
        for h, v in zip(headers, row):
            rec[h] = "" if v is None else str(v)
        result.append(rec)
    return result


def read_csv(path: Path) -> list[dict]:
    with open(str(path), newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def main() -> None:
    parser = argparse.ArgumentParser(description="Import tags into pixgrep index")
    parser.add_argument("--file", required=True, help="xlsx or csv source file")
    parser.add_argument("--sheet", default=None, help="Sheet name (xlsx only)")
    parser.add_argument("--mapping", required=True, help="JSON column-mapping file")
    args = parser.parse_args()

    mapping = json.loads(Path(args.mapping).read_text(encoding="utf-8"))
    filename_key = mapping["filename"]
    field_keys: dict[str, str] = mapping["fields"]
    text_keys: list[str] = mapping.get("text", [])

    path = Path(args.file)
    if path.suffix.lower() in (".xlsx", ".xls"):
        records = read_xlsx(path, args.sheet)
    else:
        records = read_csv(path)

    print(f"Loaded {len(records)} records from {path.name}")

    cfg = load_config()
    report = import_tags(
        cfg.index_dir,
        records,
        filename_key=filename_key,
        field_keys=field_keys,
        text_keys=text_keys,
    )
    print(report)


if __name__ == "__main__":
    main()
